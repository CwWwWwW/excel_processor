import pytest
openpyxl = pytest.importorskip("openpyxl")
from pathlib import Path
from openpyxl import Workbook, load_workbook
from contracts.job import FileSpec, JobSpec, OutputSpec
from contracts.operation import OperationSpec, TargetSpec
from scheduler.job_scheduler import JobScheduler

def test_scheduler_commits_candidate_to_output(tmp_path, monkeypatch):
    monkeypatch.setenv('EXCEL_PROCESSOR_RUNTIME', str(tmp_path/'rt'))
    src=tmp_path/'?.xlsx'; wb=Workbook(); wb.active.title='A'; wb.create_sheet('B'); wb.save(src)
    job=JobSpec(name='multi', files=(FileSpec(source_path=src),), operations=(OperationSpec(opcode='SET_VALUE', target=TargetSpec(sheet_names=('A','B'), address='A1'), parameters={'value':'ok'}),), output=OutputSpec(output_directory=tmp_path/'out'))
    result=JobScheduler(tmp_path/'rt').run(job)
    assert result.success
    out=tmp_path/'out'/'?_processed.xlsx'
    assert out.exists()
    saved=load_workbook(out)
    assert saved['A']['A1'].value == 'ok'
    assert saved['B']['A1'].value == 'ok'
    saved.close()
    original=load_workbook(src)
    assert original['A']['A1'].value is None
    original.close()

def test_candidate_validation_failure_does_not_commit(tmp_path):
    src=tmp_path/'bad.xls'; src.write_bytes(b'not excel')
    job=JobSpec(name='bad', files=(FileSpec(source_path=src),), operations=(), output=OutputSpec(output_directory=tmp_path/'out'))
    result=JobScheduler(tmp_path/'rt').run(job)
    assert not result.success
    assert not (tmp_path/'out'/'bad_processed.xls').exists()
