import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import Workbook, load_workbook

from contracts.job import AtomicityMode, ErrorPolicy, FileSpec, JobSpec, OutputSpec
from contracts.operation import OperationSpec, TargetSpec
from scheduler.job_scheduler import JobScheduler


def _book(path, title="Sheet"):
    wb = Workbook()
    wb.active.title = title
    wb.save(path)


def test_job_atomicity_one_file_failure_rolls_back_all_outputs(tmp_path):
    src_ok = tmp_path / "ok.xlsx"
    src_bad = tmp_path / "bad.xlsx"
    _book(src_ok)
    _book(src_bad)
    ok = FileSpec(source_path=src_ok)
    bad = FileSpec(source_path=src_bad)
    job = JobSpec(
        name="job-atomic",
        files=(ok, bad),
        atomicity_mode=AtomicityMode.JOB,
        operations=(
            OperationSpec(opcode="SET_VALUE", target=TargetSpec(file_ids=(ok.file_id,), sheet_names=("Sheet",), address="A1"), parameters={"value": "ok"}),
            OperationSpec(opcode="SET_VALUE", target=TargetSpec(file_ids=(bad.file_id,), sheet_names=("Sheet",)), parameters={"value": "missing-address"}, error_policy=ErrorPolicy.SKIP_FILE),
        ),
        output=OutputSpec(output_directory=tmp_path / "out"),
    )
    result = JobScheduler(tmp_path / "rt").run(job)
    assert not result.success
    assert result.status == "ROLLED_BACK"
    assert not list((tmp_path / "out").glob("*.xlsx"))


def test_file_atomicity_partial_success(tmp_path):
    src_ok = tmp_path / "ok.xlsx"
    src_bad = tmp_path / "bad.xlsx"
    _book(src_ok)
    _book(src_bad)
    ok = FileSpec(source_path=src_ok)
    bad = FileSpec(source_path=src_bad)
    job = JobSpec(
        name="file-atomic",
        files=(ok, bad),
        atomicity_mode=AtomicityMode.FILE,
        operations=(
            OperationSpec(opcode="SET_VALUE", target=TargetSpec(file_ids=(ok.file_id,), sheet_names=("Sheet",), address="A1"), parameters={"value": "ok"}),
            OperationSpec(opcode="SET_VALUE", target=TargetSpec(file_ids=(bad.file_id,), sheet_names=("Sheet",)), parameters={"value": "missing-address"}, error_policy=ErrorPolicy.SKIP_FILE),
        ),
        output=OutputSpec(output_directory=tmp_path / "out"),
    )
    result = JobScheduler(tmp_path / "rt").run(job)
    assert result.success
    assert result.status == "PARTIAL_SUCCESS"
    saved = load_workbook(tmp_path / "out" / "ok_processed.xlsx")
    assert saved["Sheet"]["A1"].value == "ok"
    saved.close()
    assert not (tmp_path / "out" / "bad_processed.xlsx").exists()
