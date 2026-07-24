import pytest
openpyxl = pytest.importorskip("openpyxl")
pandas = pytest.importorskip("pandas")
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from contracts.job import FileSpec, JobSpec, OutputSpec
from contracts.operation import OperationSpec, TargetSpec
from scheduler.job_scheduler import JobScheduler

def test_dataframe_dedup_does_not_replace_sheet(tmp_path):
    src=tmp_path/'data.xlsx'; wb=Workbook(); ws=wb.active; ws.title='Data'; ws['A1']='id'; ws['A1'].fill=PatternFill('solid', fgColor='FFFF00'); ws.append([1]); ws.append([1]); wb.save(src)
    job=JobSpec(name='dedup', files=(FileSpec(source_path=src),), operations=(OperationSpec(opcode='DEDUP_DATA', target=TargetSpec(sheet_names=('Data',)), parameters={'subset':['id']}),), output=OutputSpec(output_directory=tmp_path/'out', allow_warnings=True))
    result=JobScheduler(tmp_path/'rt').run(job)
    assert result.success
    out=load_workbook(tmp_path/'out'/'data_processed.xlsx')
    assert out['Data']['A1'].fill.fgColor.rgb.endswith('FFFF00')
    assert out['Data'].max_row == 2
    out.close()
