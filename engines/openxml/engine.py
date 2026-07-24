from __future__ import annotations
import os, shutil, time, zipfile
from pathlib import Path
from uuid import uuid4
from contracts.job import EngineMode
from contracts.plan import ExecutionContext, OperationCommand
from contracts.result import ChangeRecord, OperationResult
OOXML_EXTS={'.xlsx','.xlsm','.xltx','.xltm'}
def _keep_vba(path: Path) -> bool: return path.suffix.lower() in {'.xlsm','.xltm'}
def _safe_save_workbook(wb, target: Path) -> None:
    temp = target.with_name(f'.{target.name}.{uuid4().hex}.tmp{target.suffix}')
    try:
        wb.save(temp); wb.close()
        with zipfile.ZipFile(temp) as zf:
            bad=zf.testzip()
            if bad: raise ValueError(f'OOXML ZIP ???{bad}')
        os.replace(temp, target)
    finally:
        if temp.exists():
            try: temp.unlink()
            except OSError as exc:
                _ = exc
class OpenXmlEngine:
    name='openxml'
    def execute(self, command: OperationCommand, context: ExecutionContext) -> OperationResult:
        started=time.perf_counter(); op=command.operation; wb=None
        try:
            if command.working_path.suffix.lower() not in OOXML_EXTS: raise ValueError('OpenXML ????? OOXML ???')
            from openpyxl import load_workbook
            if op.opcode=='SET_VALUE':
                wb=load_workbook(command.working_path, keep_vba=_keep_vba(command.working_path)); ws=wb[command.resolved_sheet or (op.target.sheet_names or (wb.sheetnames[0],))[0]]; addr=op.target.address or op.parameters.get('address')
                if not addr: raise ValueError('SET_VALUE ?? target.address ? parameters.address')
                old=ws[addr].value; new=op.parameters.get('value'); ws[addr]=new; title=ws.title; _safe_save_workbook(wb, command.working_path); wb=None
                return OperationResult(operation_id=op.operation_id, file_id=command.file_id, success=True, engine_used=EngineMode.OPENXML, affected_objects=1, changes=(ChangeRecord(file_id=command.file_id, sheet_name=title, address=addr, object_type='cell', field_name='value', old_value=old, new_value=new),), duration_ms=int((time.perf_counter()-started)*1000))
            if op.opcode=='DELETE_ROWS':
                wb=load_workbook(command.working_path, keep_vba=_keep_vba(command.working_path)); ws=wb[command.resolved_sheet or (op.target.sheet_names or (wb.sheetnames[0],))[0]]; rows=op.parameters.get('rows')
                if rows: row_list=sorted({int(r) for r in rows}, reverse=True); amount=len(row_list); [ws.delete_rows(r,1) for r in row_list]
                else: idx=int(op.parameters.get('idx', op.parameters.get('row',1))); amount=int(op.parameters.get('amount',1)); [ws.delete_rows(r,1) for r in range(idx+amount-1, idx-1, -1)]
                _safe_save_workbook(wb, command.working_path); wb=None
                return OperationResult(operation_id=op.operation_id, file_id=command.file_id, success=True, engine_used=EngineMode.OPENXML, affected_objects=amount, duration_ms=int((time.perf_counter()-started)*1000))
            if op.opcode=='COPY_TO_CANDIDATE' and command.output_path:
                shutil.copy2(command.working_path, command.output_path)
                return OperationResult(operation_id=op.operation_id, file_id=command.file_id, success=True, engine_used=EngineMode.OPENXML, affected_objects=1, duration_ms=int((time.perf_counter()-started)*1000))
            if op.opcode=='SAVE_AS': raise ValueError('???? SAVE_AS ???? Excel COM')
            raise ValueError(f'OpenXML ????????{op.opcode}')
        except Exception as exc:
            if wb is not None:
                try: wb.close()
                except Exception as close_exc:
                    _ = close_exc
            return OperationResult(operation_id=op.operation_id, file_id=command.file_id, success=False, engine_used=EngineMode.OPENXML, errors=(str(exc),), duration_ms=int((time.perf_counter()-started)*1000))
