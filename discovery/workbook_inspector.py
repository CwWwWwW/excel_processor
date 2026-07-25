from __future__ import annotations
import hashlib
from pathlib import Path
from uuid import UUID
from pydantic import BaseModel, ConfigDict, Field, field_serializer


class WorkbookSnapshot(BaseModel):
    model_config = ConfigDict(frozen=True)
    schema_version: str = "1.0"
    file_id: UUID
    source_path: Path
    sha256: str
    sheets: tuple[str, ...] = ()
    file_format: str = ""
    has_vba: bool = False
    has_tables: bool = False
    has_charts: bool = False
    has_pivot_tables: bool = False
    has_data_model: bool = False
    has_power_query: bool = False
    has_external_connections: bool = False
    has_activex: bool = False
    has_ole_objects: bool = False
    has_conditional_formatting: bool = False
    has_data_validation: bool = False
    has_defined_names: bool = False
    has_external_links: bool = False
    object_counts: dict[str, int] = Field(default_factory=dict)
    warnings: tuple[str, ...] = ()

    @field_serializer("source_path")
    def _p(self, v: Path) -> str:
        return str(v)

    def stable_hash(self) -> str:
        return hashlib.sha256(self.model_dump_json().encode("utf-8")).hexdigest()


def inspect_workbook(file_id: UUID, path: Path) -> WorkbookSnapshot:
    from .file_scanner import has_vba_project, sha256_file

    warnings: list[str] = []
    sheets: tuple[str, ...] = ()
    counts: dict[str, int] = {}
    flags: dict[str, bool] = {}
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(path, read_only=False, keep_vba=True, data_only=False)
            sheets = tuple(wb.sheetnames)
            counts["worksheets"] = len(sheets)
            flags["has_tables"] = any(bool(getattr(ws, "tables", {})) for ws in wb.worksheets)
            flags["has_charts"] = any(bool(getattr(ws, "_charts", ())) for ws in wb.worksheets)
            flags["has_conditional_formatting"] = any(bool(getattr(ws, "conditional_formatting", ())) for ws in wb.worksheets)
            flags["has_data_validation"] = any(bool(getattr(getattr(ws, "data_validations", None), "dataValidation", ())) for ws in wb.worksheets)
            flags["has_defined_names"] = bool(list(getattr(wb, "defined_names", [])))
            flags["has_external_links"] = bool(getattr(wb, "_external_links", ()))
            wb.close()
        except Exception as exc:
            warnings.append(f"OpenXML snapshot read failed: {exc}")
    else:
        warnings.append("This format needs Excel COM for a complete workbook snapshot")

    return WorkbookSnapshot(
        file_id=file_id,
        source_path=path,
        sha256=sha256_file(path),
        sheets=sheets,
        file_format=suffix.lstrip("."),
        has_vba=has_vba_project(path),
        object_counts=counts,
        warnings=tuple(warnings),
        **flags,
    )
